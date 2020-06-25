from SharedMethods import get_dict_of_paths
import SimpleITK as sitk
import os
import re
from openpyxl import load_workbook
import numpy as np


# returns dice coefficent, mean overlap and volume similarity measurements
def calculate_label_overlap_measures(gt_img_path, pred_img_path):
    # load images as int since float is not supported by GetDiceCoefficient()
    gt_img = sitk.ReadImage(gt_img_path, sitk.sitkInt32)
    pred_img = sitk.ReadImage(pred_img_path, sitk.sitkInt32)

    # calculate dice
    dice_filter = sitk.LabelOverlapMeasuresImageFilter()
    dice_filter.Execute(gt_img, pred_img)
    dice = dice_filter.GetDiceCoefficient()
    #mean_overlap = dice_filter.GetMeanOverlap()
    #volume_similarity = dice_filter.GetVolumeSimilarity()

    print("dice coefficient {}".format(dice))
    #print("mean overlap {}".format(mean_overlap))
    #print("volume similarity {}".format(volume_similarity))

    return dice


# returns hausdorff distance and average hausdorff distance measurements
def calculate_hausdorff_distance(gt_img_path, pred_img_path):
    # load images
    gt_img = sitk.ReadImage(gt_img_path)
    pred_img = sitk.ReadImage(pred_img_path)

    # calculate hausdorff
    hd_filter = sitk.HausdorffDistanceImageFilter()
    hd_filter.Execute(gt_img, pred_img)
    avg_hd = hd_filter.GetAverageHausdorffDistance()
    hd = hd_filter.GetHausdorffDistance()

    print("hausdorff distance {}".format(hd))
    print("average hausdorff distance {}".format(avg_hd))

    return hd, avg_hd


#TODO: takes only 1 image
def calculate_danielsson_distance(gt_img_path, pred_img_path):
    # load images
    gt_img = sitk.ReadImage(gt_img_path)
    pred_img = sitk.ReadImage(pred_img_path)

    # calculate hausdorff
    dd_filter = sitk.DanielssonDistanceMapImageFilter()
    dd_filter.SetSquaredDistance(False)
    dd_filter.Execute(gt_img, pred_img)
    dd = dd_filter.GetSquaredDistance()

    print("squared danielsson distance {}".format(dd))

    return dd


def calculate_mean(results):
    # organize metrics of all patients
    hd = []
    avg_hd = []
    dice = []
    #avg_ol = []
    #vs = []
    for result in results:
        hd.append(result[1])
        avg_hd.append(result[2])
        dice.append((result[3]))
        #avg_ol.append(result[4])
        #vs.append(result[5])

    # calculate mean of all metrics
    mean_hd = np.mean(hd)
    mean_avg_hd = np.mean(avg_hd)
    mean_dice = np.mean(dice)
    #mean_avg_ol = np.mean(avg_ol)
    #mean_vs = np.mean(vs)

    mean = ["mean", mean_hd, mean_avg_hd, mean_dice]
    return mean


def calculate_standard_dv(results):
    # organize metrics of all patients
    hd = []
    avg_hd = []
    dice = []
    #avg_ol = []
    #vs = []
    for result in results:
        hd.append(result[1])
        avg_hd.append(result[2])
        dice.append((result[3]))
        #avg_ol.append(result[4])
        #vs.append(result[5])

    # calculate std of all metrics
    std_hd = np.std(hd)
    std_avg_hd = np.std(avg_hd)
    std_dice = np.std(dice)
    #std_avg_ol = np.std(avg_ol)
    #std_vs = np.std(vs)

    std = ["standard d", std_hd, std_avg_hd, std_dice]
    return std


def evaluate_predictions(pred_path, gt_path):
    results = []

    dict_gt_paths = get_dict_of_paths(gt_path)

    # evaluate every prediction
    for prediction in os.scandir(pred_path):
        # find patient number in file name
        regex = re.compile(r'\d+')
        patient_no = int(regex.search(prediction.name).group(0))
        gt_img_path = dict_gt_paths[patient_no]

        print("Patient Number : {}".format(patient_no))
        hd, avg_hd = calculate_hausdorff_distance(gt_img_path, prediction.path)
        dice = calculate_label_overlap_measures(gt_img_path, prediction.path)
        #danielsson = calculate_danielsson_distance(gt_img_path, prediction.path)

        print("")

        results.append([patient_no, hd, avg_hd, dice])

    mean = calculate_mean(results)
    std = calculate_standard_dv(results)

    return results, mean, std


'''
https://realpython.com/openpyxl-excel-spreadsheets-python/
'''


def evaluate(SAVE_PATH, ORGAN):
    # open excel sheet
    wb = load_workbook(filename="{}Evaluation {}.xlsx".format(SAVE_PATH, ORGAN))
    sheet = wb.active

    # get metrics
    path_results_orig = "{}results/orig/".format(SAVE_PATH)
    path_y_test_orig = "{}ytest/orig/".format(SAVE_PATH)
    results, mean, std = evaluate_predictions(path_results_orig, path_y_test_orig)

    # write metrics into excel
    number_of_results = len(results)
    number_of_metrics = len(results[0])
    start_row = 4
    start_col = 1
    for i in range(0, number_of_results):
        for j in range(0, number_of_metrics):
            row = start_row + i  # start a new row for every result
            col = start_col + j  # put metrcis into the right column
            # cut all values to 2 decimal places. Except patient number
            metric_value = "{:.2f}".format(results[i][j])
            if j == 0:
                metric_value = "{}".format(results[i][j])
            sheet.cell(column=col, row=row, value=metric_value)

    # write mean and std also into excel
    mean_row = start_row + number_of_results
    std_row = mean_row + 1
    for i in range(0, len(mean)):
        col = start_col + i
        if i == 0:
            mean_value = "{}".format(mean[i])
            std_value = "{}".format(std[i])
        else:
            mean_value = "{:.2f}".format(mean[i])
            std_value = "{:.2f}".format(std[i])
        sheet.cell(column=col, row=mean_row, value=mean_value)
        sheet.cell(column=col, row=std_row, value=std_value)

    wb.save("{}Evaluation {}.xlsx".format(SAVE_PATH, ORGAN))