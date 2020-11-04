import numpy as np
from sklearn.model_selection import KFold
from SharedMethods import find_patient_no_in_file_name
import os
import tensorflow as tf
from UNet3D.Prepare import prepare
from UNet3D.Train import train
from UNet3D.Apply import apply
from UNet3D.Evaluate import evaluate, summarize_eval
from UNet2D.Prepare2D import create_excel_sheet2D
from UNet2D.Train2D import train2D
from UNet2D.Apply2D import apply2D
from UNet2D.Evaluate2D import evaluate2D
import time

from openpyxl import load_workbook
import numpy as np
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl import Workbook
import openpyxl as op

def get_files_in_path(path):
    # count how many files are in SCAN_PATH
    all_patient_numbers = set()
    for file in os.scandir(path):
        patient_no = find_patient_no_in_file_name(file.name)
        all_patient_numbers.add(patient_no)
    amount_patients = len(all_patient_numbers)
    all_patient_numbers_arr = np.zeros(amount_patients)
    i = 0
    for val in all_patient_numbers:
        all_patient_numbers_arr[i] = val
        i = i + 1

    return all_patient_numbers_arr, amount_patients

#TODO
def write_kfold_into_file(parts):
    number = 0
    for train_set, test_set in parts:
        print("#{} train set: {} files".format(number, len(train_set)))
        print(train_set)
        print("#{} test set: {} files".format(number, len(test_set)))
        print(test_set)
        print("")
        number = number + 1

def run_KfoldCV(SCAN_PATH, GT_BB_PATH, RRF_BB_PATH, GT_SEG_PATH, SAVE_PATH, DIMENSIONS, BATCH, EPOCHS, organs):
    kfold = KFold(5, True, 1)
    data, amount_patients = get_files_in_path(SCAN_PATH)
    parts = kfold.split(data)
    write_kfold_into_file(parts)
    exit()
    number = 0
    for train_set, test_set in parts:
        print("train set {}".format(len(train_set)))
        print(train_set)
        print("test set {}".format(len(test_set)))
        print(test_set)
        exit()
        number = number + 1
        for organ in organs:
            if organ == 'pancreas':
                thresh = 0.3
            else:
                thresh = 0.5

            prepare(SCAN_PATH, GT_BB_PATH, RRF_BB_PATH, GT_SEG_PATH, SAVE_PATH, DIMENSIONS, organ, train_set, test_set)
            create_excel_sheet2D(SAVE_PATH, organ, train_set, test_set)

            # 3D
            start = time.time()
            train(SAVE_PATH, DIMENSIONS, organ, 0.0, BATCH, EPOCHS)
            end = time.time()
            elapsed_time = end - start

            apply(SCAN_PATH, RRF_BB_PATH, SAVE_PATH, DIMENSIONS, organ, thresh)
            evaluate(SAVE_PATH, organ, number, elapsed_time)

            # 2D
            start = time.time()
            train2D(SAVE_PATH, DIMENSIONS, organ, 0.0, BATCH, EPOCHS)
            end = time.time()
            elapsed_time = end - start

            apply2D(SCAN_PATH, RRF_BB_PATH, SAVE_PATH, DIMENSIONS, organ, thresh)
            evaluate2D(SAVE_PATH, organ, number, elapsed_time)


#TODO
def summarize_eval(SAVE_PATH, ORGAN):
    # create excel sheet
    eval_wb = Workbook()
    eval_sheet = eval_wb.active
    eval_sheet.title = "summarize eval"

    # create headings and apply style
    headings_style = NamedStyle(
        name="daria",
        font=Font(color='000000', bold=True),
        alignment=Alignment(horizontal='left')
    )
    headings_row = '1'
    headings = ["file #", "mean dice coeff",
                "standard d."]
    eval_sheet.append(headings)
    for cell in eval_sheet[headings_row]:
        cell.style = headings_style

    # make cells wider
    eval_sheet.column_dimensions['A'].width = 50
    eval_sheet.column_dimensions['B'].width = 20
    eval_sheet.column_dimensions['C'].width = 20

    mean_dice_row = 24
    mean_standard_d_row = 25
    relevant_col = 4

    curr_row = 2
    curr_col = 1

    path = "{}eval/".format(SAVE_PATH)
    for file in os.scandir(path):
        found = file.name.find(ORGAN)
        if found is not -1:
            # open
            wb_obj = op.load_workbook(file)
            # get active sheet
            sheet_obj = wb_obj.active
            # read cell
            cell_mean_dice = sheet_obj.cell(row=mean_dice_row, column=relevant_col)
            cell_standard_d = sheet_obj.cell(row=mean_standard_d_row, column=relevant_col)
            # get value
            mean_dice = cell_mean_dice.value
            standard_d = cell_standard_d.value

            # write into evaluation summary sheet
            eval_sheet.cell(column=curr_col, row=curr_row, value=file.name)  # name
            eval_sheet.cell(column=curr_col+1, row=curr_row, value=mean_dice)
            eval_sheet.cell(column=curr_col+2, row=curr_row, value=standard_d)

            curr_row = curr_row+1

    eval_wb.save("{}Evaluation Summary {}.xlsx".format(SAVE_PATH, ORGAN))