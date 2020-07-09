import os
import re
import numpy
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl import Workbook, load_workbook


# read bounding box coordinates
def get_bb_coordinates(box_path):
    # open vtk file and get coordinates
    bb_file = open(box_path, 'r')
    lines = bb_file.readlines()

    # get coordinates
    numbers1 = lines[6].split()
    x0 = float(numbers1[0])
    y0 = float(numbers1[1])
    z0 = float(numbers1[2])

    # get coordinates
    numbers2 = lines[12].split()
    x1 = float(numbers2[0])
    y1 = float(numbers2[1])
    z1 = float(numbers2[2])

    # close file
    bb_file.close()

    # add coordinates to array
    bb_coords = [x0, x1, y0, y1, z0, z1]

    return bb_coords


# given the name of an organ it returns the organs label number
def get_organ_label(organ):
    # define dictionary (to simulate switch-case)
    switcher = {
        "liver": 170,
        "left_kidney": 156,
        "right_kidney": 157,
        "spleen": 160,
        "pancreas": 150
    }

    # if given organ isn't a defined key, return "no valid organ"
    organ_label = switcher.get(organ, "no valid organ")

    # raise error message if no valid organ name was given
    if organ_label == "no valid organ":
        raise ValueError("'{}' is no valid organ name. Valid names are: "
                         "'liver', 'left_kidney', 'right_kidney', 'spleen', "
                         "'pancreas'".format(organ))
    else:
        return organ_label


# load all paths to all .vtk files in a folder into a list that is sorted by ascending patient numbers
# but only use the .vtk files of a given organ
# assuming files contain patient numbers anywhere in the filename
# assuming files contain organ numbers followed by "_bb" anywhere in the filename
def get_dict_of_paths(path, organ=None):
    # if an organ was given, check if name is valid and get label for organ
    if organ is not None:
        organ_label = get_organ_label(organ)
        organ_label = "{}_bb".format(organ_label)

    dict_of_paths = {}

    # go through every file in directory
    # (this isn't done in a organized way, files seem to be accessed rather randomly)
    for file in os.scandir(path):
        # find patient number in file name
        regex = re.compile(r'\d+')
        patient_no = int(regex.search(file.name).group(0))

        if organ is not None:
            # write filepath (to file that contains the organ label) into dictionary with patient number as key
            if organ_label in file.name:
                dict_of_paths[patient_no] = file.path
        else:
            # write filepath into dictionary with patient number as key
            dict_of_paths[patient_no] = file.path

    return dict_of_paths


def get_bb_differences(gt_bb_dict, bb_dict):
    result_dict = {}

    # bbs in ascending order
    for key in sorted(gt_bb_dict.keys()):
        gt_bb_path = gt_bb_dict[key]
        bb_path = bb_dict[key]

        # get gt bounding box coordinates
        gt_bb_coords = get_bb_coordinates(gt_bb_path)
        # width
        x0_gt_bb = gt_bb_coords[0]
        x1_gt_bb = gt_bb_coords[1]
        # height
        y0_gt_bb = gt_bb_coords[2]
        y1_gt_bb = gt_bb_coords[3]
        # depth
        z0_gt_bb = gt_bb_coords[4]
        z1_gt_bb = gt_bb_coords[5]
        # print("gt bb coords: {} ".format(gt_bb_coords))

        # get bounding box coordinates
        bb_coords = get_bb_coordinates(bb_path)
        # width
        x0_bb = bb_coords[0]
        x1_bb = bb_coords[1]
        # height
        y0_bb = bb_coords[2]
        y1_bb = bb_coords[3]
        # depth
        z0_bb = bb_coords[4]
        z1_bb = bb_coords[5]
        # print("bb coords: {} ".format(bb_coords))

        # calculate width diff
        x0 = abs(x0_gt_bb - x0_bb)
        x1 = abs(x1_gt_bb - x1_bb)
        # calculate height diff
        y0 = abs(y0_gt_bb - y0_bb)
        y1 = abs(y1_gt_bb - y1_bb)
        # calculate depth diff
        z0 = abs(z0_gt_bb - z0_bb)
        z1 = abs(z1_gt_bb - z1_bb)
        # print("differences: x0: {}, x1: {}, y0: {}, y1: {}, z0: {}, z1: {}".format(x0, x1, y0, y1, z0, z1))
        # print("")

        arr = [x0, x1, y0, y1, z0, z1]
        result_dict[key] = arr

    return result_dict


def organize_values(dict_bb_differences):
    x0_arr = []
    x1_arr = []
    y0_arr = []
    y1_arr = []
    z0_arr = []
    z1_arr = []

    # put x0 values of every patient in one array. Same with x1,y0,y1,z0,z1
    for key in dict_bb_differences.keys():
        curr_arr = dict_bb_differences[key]
        x0_arr.append(curr_arr[0])
        x1_arr.append(curr_arr[1])
        y0_arr.append(curr_arr[2])
        y1_arr.append(curr_arr[3])
        z0_arr.append(curr_arr[4])
        z1_arr.append(curr_arr[5])

    result_arr = [x0_arr, x1_arr, y0_arr, y1_arr, z0_arr, z1_arr]
    return result_arr


def calculate_mean(organized_values):
    mean_x0 = numpy.mean(organized_values[0])
    mean_x1 = numpy.mean(organized_values[1])
    mean_y0 = numpy.mean(organized_values[2])
    mean_y1 = numpy.mean(organized_values[3])
    mean_z0 = numpy.mean(organized_values[4])
    mean_z1 = numpy.mean(organized_values[5])

    result_arr = [mean_x0, mean_x1, mean_y0, mean_y1, mean_z0, mean_z1]

    return result_arr


def calculate_standard_deviation(organized_values):
    std_x0 = numpy.std(organized_values[0])
    std_x1 = numpy.std(organized_values[1])
    std_y0 = numpy.std(organized_values[2])
    std_y1 = numpy.std(organized_values[3])
    std_z0 = numpy.std(organized_values[4])
    std_z1 = numpy.std(organized_values[5])

    result_arr = [std_x0, std_x1, std_y0, std_y1, std_z0, std_z1]

    return result_arr


# compare walls separately
def compare(gt_bb_path, bb_path, organ):
    gt_bb_dict = get_dict_of_paths(gt_bb_path, organ)
    bb_dict = get_dict_of_paths(bb_path, organ)

    dict_bb_differences = get_bb_differences(gt_bb_dict, bb_dict)

    organized_values = organize_values(dict_bb_differences)
    mean = calculate_mean(organized_values)
    std = calculate_standard_deviation(organized_values)

    print("CALCULATE MEAN AND STD FOR ALL WALLS FOR {} SEPARATELY". format(organ))
    print(" mean x0,x1,y0,y1,z0,z1:")
    print(" {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}".format(mean[0], mean[1], mean[2], mean[3], mean[4], mean[5]))
    print(" std x0,x1,y0,y1,z0,z1:")
    print(" {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}".format(std[0], std[1], std[2], std[3], std[4], std[5]))
    print(" mean mean: {:.2f}".format(numpy.mean(mean)))
    print(" mean std: {:.2f}".format(numpy.mean(std)))
    print("")

    return dict_bb_differences, mean, std


# compare using all values at once
def compare2(gt_bb_path, bb_path, organ):
    gt_bb_dict = get_dict_of_paths(gt_bb_path, organ)
    bb_dict = get_dict_of_paths(bb_path, organ)

    dict_bb_differences = get_bb_differences(gt_bb_dict, bb_dict)

    # put all values (x0, x1,y0,y1,z0,z1) of every patient in one array
    all_differences_arr = []
    for key in dict_bb_differences.keys():
        curr_arr = dict_bb_differences[key]
        all_differences_arr.append(curr_arr[0])
        all_differences_arr.append(curr_arr[1])
        all_differences_arr.append(curr_arr[2])
        all_differences_arr.append(curr_arr[3])
        all_differences_arr.append(curr_arr[4])
        all_differences_arr.append(curr_arr[5])

    mean = numpy.mean(all_differences_arr)
    std = numpy.std(all_differences_arr)

    print("calculate mean and std for all walls for {} at once".format(organ))
    print(" mean: {:.2f}".format(mean))
    print(" std: {:.2f}".format(std))
    print("")
    print("")
    return mean, std


def create_excel_sheet(SAVE_PATH, ORGAN):
    # create excel sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = ORGAN

    # create headings and apply style
    headings_style = NamedStyle(
        name="daria",
        font=Font(color='000000', bold=True),
        alignment=Alignment(horizontal='left')
    )
    headings1 = ["", "L", "R", "A", "P", "I", "S", "Ø"]
    #headings1 = ["", "x0", "x1", "y0", "y1", "z0", "z1", "Ø"]
    row = 2
    while row < 26:
        for i in range(0, len(headings1)):
            col = i+1
            sheet.cell(column=col, row=row, value=headings1[i]).style = headings_style
        row = row + 5

    col = 1
    content = ["LIVER", "", "mean", "std", "",
               "L KIDNEY", "", "mean", "std", "",
               "R KIDNEY", "", "mean", "std", "",
               "SPLEEN", "", "mean", "std", "",
               "PANCREAS", "", "mean", "std", ""]
    for i in range(0,len(content)-1):
        row = i+1
        sheet.cell(column=col, row=row, value=content[i]).style = headings_style

    # make cells wider
    column_width = 15
    sheet.column_dimensions['A'].width = column_width
    sheet.column_dimensions['B'].width = column_width
    sheet.column_dimensions['C'].width = column_width
    sheet.column_dimensions['D'].width = column_width
    sheet.column_dimensions['E'].width = column_width
    sheet.column_dimensions['F'].width = column_width

    wb.save("{}RRF BB Evaluation.xlsx".format(SAVE_PATH))


def write_into_sheet(content, row, SAVE_PATH):
    # open excel sheet
    wb = load_workbook(filename="{}RRF BB Evaluation.xlsx".format(SAVE_PATH))
    sheet = wb.active
    for i in range(0, len(content)):
        col = i+2
        value = "{:.2f}".format(content[i])
        sheet.cell(column=col, row=row, value=value)

    wb.save("{}RRF BB Evaluation.xlsx".format(SAVE_PATH))

