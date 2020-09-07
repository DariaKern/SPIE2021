from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl import Workbook, load_workbook
import numpy
from DATA.SharedMethods_old import get_dict_of_paths, get_bb_coordinates


def get_bb_differences(gt_bb_dict, bb_dict):
    result_dict = {}

    # bbs in ascending order
    for key in sorted(gt_bb_dict.keys()):
        gt_bb_path = gt_bb_dict[key]
        bb_path = bb_dict[key]

        # get gt bounding box coordinates
        gt_bb_coords = get_bb_coordinates(gt_bb_path)
        # width
        x0_gt_bb = gt_bb_coords[0]
        x1_gt_bb = gt_bb_coords[1]
        # height
        y0_gt_bb = gt_bb_coords[2]
        y1_gt_bb = gt_bb_coords[3]
        # depth
        z0_gt_bb = gt_bb_coords[4]
        z1_gt_bb = gt_bb_coords[5]
        # print("gt bb coords: {} ".format(gt_bb_coords))

        # get bounding box coordinates
        bb_coords = get_bb_coordinates(bb_path)
        # width
        x0_bb = bb_coords[0]
        x1_bb = bb_coords[1]
        # height
        y0_bb = bb_coords[2]
        y1_bb = bb_coords[3]
        # depth
        z0_bb = bb_coords[4]
        z1_bb = bb_coords[5]
        # print("bb coords: {} ".format(bb_coords))

        # calculate width diff
        x0 = abs(x0_gt_bb - x0_bb)
        x1 = abs(x1_gt_bb - x1_bb)
        # calculate height diff
        y0 = abs(y0_gt_bb - y0_bb)
        y1 = abs(y1_gt_bb - y1_bb)
        # calculate depth diff
        z0 = abs(z0_gt_bb - z0_bb)
        z1 = abs(z1_gt_bb - z1_bb)
        # print("differences: x0: {}, x1: {}, y0: {}, y1: {}, z0: {}, z1: {}".format(x0, x1, y0, y1, z0, z1))
        # print("")

        arr = [x0, x1, y0, y1, z0, z1]
        result_dict[key] = arr

    return result_dict


def organize_values(dict_bb_differences):
    x0_arr = []
    x1_arr = []
    y0_arr = []
    y1_arr = []
    z0_arr = []
    z1_arr = []

    # put x0 values of every patient in one array. Same with x1,y0,y1,z0,z1
    for key in dict_bb_differences.keys():
        curr_arr = dict_bb_differences[key]
        x0_arr.append(curr_arr[0])
        x1_arr.append(curr_arr[1])
        y0_arr.append(curr_arr[2])
        y1_arr.append(curr_arr[3])
        z0_arr.append(curr_arr[4])
        z1_arr.append(curr_arr[5])

    result_arr = [x0_arr, x1_arr, y0_arr, y1_arr, z0_arr, z1_arr]
    return result_arr


def calculate_mean(organized_values):
    mean_x0 = numpy.mean(organized_values[0])
    mean_x1 = numpy.mean(organized_values[1])
    mean_y0 = numpy.mean(organized_values[2])
    mean_y1 = numpy.mean(organized_values[3])
    mean_z0 = numpy.mean(organized_values[4])
    mean_z1 = numpy.mean(organized_values[5])

    result_arr = [mean_x0, mean_x1, mean_y0, mean_y1, mean_z0, mean_z1]

    return result_arr


def calculate_standard_deviation(organized_values):
    std_x0 = numpy.std(organized_values[0])
    std_x1 = numpy.std(organized_values[1])
    std_y0 = numpy.std(organized_values[2])
    std_y1 = numpy.std(organized_values[3])
    std_z0 = numpy.std(organized_values[4])
    std_z1 = numpy.std(organized_values[5])

    result_arr = [std_x0, std_x1, std_y0, std_y1, std_z0, std_z1]

    return result_arr


# compare walls separately
def compare(gt_bb_path, bb_path, organ):
    gt_bb_dict = get_dict_of_paths(gt_bb_path, organ)
    bb_dict = get_dict_of_paths(bb_path, organ)

    dict_bb_differences = get_bb_differences(gt_bb_dict, bb_dict)

    organized_values = organize_values(dict_bb_differences)
    mean = calculate_mean(organized_values)
    std = calculate_standard_deviation(organized_values)

    print("CALCULATE MEAN AND STD FOR ALL WALLS FOR {} SEPARATELY". format(organ))
    print(" mean x0,x1,y0,y1,z0,z1:")
    print(" {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}".format(mean[0], mean[1], mean[2], mean[3], mean[4], mean[5]))
    print(" std x0,x1,y0,y1,z0,z1:")
    print(" {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}, {:.2f}".format(std[0], std[1], std[2], std[3], std[4], std[5]))
    print(" mean mean: {:.2f}".format(numpy.mean(mean)))
    print(" mean std: {:.2f}".format(numpy.mean(std)))
    print("")

    write_into_sheet(liver_mean, 3, SAVE_PATH)
    write_into_sheet(liver_std, 4, SAVE_PATH)
    
    return dict_bb_differences, mean, std


# compare using all values at once
def compare2(gt_bb_path, bb_path, organ):
    gt_bb_dict = get_dict_of_paths(gt_bb_path, organ)
    bb_dict = get_dict_of_paths(bb_path, organ)

    dict_bb_differences = get_bb_differences(gt_bb_dict, bb_dict)

    # put all values (x0, x1,y0,y1,z0,z1) of every patient in one array
    all_differences_arr = []
    for key in dict_bb_differences.keys():
        curr_arr = dict_bb_differences[key]
        all_differences_arr.append(curr_arr[0])
        all_differences_arr.append(curr_arr[1])
        all_differences_arr.append(curr_arr[2])
        all_differences_arr.append(curr_arr[3])
        all_differences_arr.append(curr_arr[4])
        all_differences_arr.append(curr_arr[5])

    mean = numpy.mean(all_differences_arr)
    std = numpy.std(all_differences_arr)

    print("calculate mean and std for all walls for {} at once".format(organ))
    print(" mean: {:.2f}".format(mean))
    print(" std: {:.2f}".format(std))
    print("")
    print("")
    return mean, std


def write_into_sheet(content, row, SAVE_PATH):
    # open excel sheet
    wb = load_workbook(filename="{}RRF BB Evaluation.xlsx".format(SAVE_PATH))
    sheet = wb.active
    for i in range(0, len(content)):
        col = i+2
        value = "{:.2f}".format(content[i])
        sheet.cell(column=col, row=row, value=value)

    wb.save("{}RRF BB Evaluation.xlsx".format(SAVE_PATH))


def create_excel_sheet(SAVE_PATH, ORGAN):
    # create excel sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = ORGAN

    # create headings and apply style
    headings_style = NamedStyle(
        name="daria",
        font=Font(color='000000', bold=True),
        alignment=Alignment(horizontal='left')
    )
    headings1 = ["", "L", "R", "A", "P", "I", "S", "Ø"]
    #headings1 = ["", "x0", "x1", "y0", "y1", "z0", "z1", "Ø"]
    row = 2
    while row < 26:
        for i in range(0, len(headings1)):
            col = i+1
            sheet.cell(column=col, row=row, value=headings1[i]).style = headings_style
        row = row + 5

    col = 1
    content = ["LIVER", "", "mean", "std", "",
               "L KIDNEY", "", "mean", "std", "",
               "R KIDNEY", "", "mean", "std", "",
               "SPLEEN", "", "mean", "std", "",
               "PANCREAS", "", "mean", "std", ""]
    for i in range(0,len(content)-1):
        row = i+1
        sheet.cell(column=col, row=row, value=content[i]).style = headings_style

    # make cells wider
    column_width = 15
    sheet.column_dimensions['A'].width = column_width
    sheet.column_dimensions['B'].width = column_width
    sheet.column_dimensions['C'].width = column_width
    sheet.column_dimensions['D'].width = column_width
    sheet.column_dimensions['E'].width = column_width
    sheet.column_dimensions['F'].width = column_width

    wb.save("{}RRF BB Evaluation.xlsx".format(SAVE_PATH))

def evaluate(GT_BB_PATH, BB_PATH, SAVE_PATH):
    create_excel_sheet(SAVE_PATH, "liver")

    # compare liver bbs
    dict_bb_diff_liver, liver_mean, liver_std = compare(GT_BB_PATH, BB_PATH, "liver")

    liver_mean_all_dimensions = compare2(GT_BB_PATH, BB_PATH, "liver")

    # compare left kidney bbs
    dict_bb_diff_l_kidney, l_kidney_mean, l_kidney_std = compare(GT_BB_PATH, BB_PATH, "left_kidney")
    write_into_sheet(l_kidney_mean, 8, SAVE_PATH)
    write_into_sheet(l_kidney_std, 9, SAVE_PATH)
    l_kidney_mean_all_dimensions = compare2(GT_BB_PATH, BB_PATH, "left_kidney")

    # compare right kidney bbs
    dict_bb_diff_r_kidney, r_kidney_mean, r_kidney_std = compare(GT_BB_PATH, BB_PATH, "right_kidney")
    write_into_sheet(r_kidney_mean, 13, SAVE_PATH)
    write_into_sheet(r_kidney_std, 14, SAVE_PATH)
    r_kidney_mean_all_dimensions = compare2(GT_BB_PATH, BB_PATH, "right_kidney")

    # compare spleen bbs
    dict_bb_diff_spleen, spleen_mean, spleen_std = compare(GT_BB_PATH, BB_PATH, "spleen")
    write_into_sheet(spleen_mean, 18, SAVE_PATH)
    write_into_sheet(spleen_std, 19, SAVE_PATH)
    spleen_mean_all_dimensions = compare2(GT_BB_PATH, BB_PATH, "spleen")

    # compare pancreas bbs
    dict_bb_diff_pancreas, pancreas_mean, pancreas_std = compare(GT_BB_PATH, BB_PATH, "pancreas")
    write_into_sheet(pancreas_mean, 23, SAVE_PATH)
    write_into_sheet(pancreas_std, 24, SAVE_PATH)
    pancreas_mean_all_dimensions = compare2(GT_BB_PATH, BB_PATH, "pancreas")
