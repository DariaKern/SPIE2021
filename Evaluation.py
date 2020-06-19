'''
Metrics for medical image processing evaluation:
https://www.ncbi.nlm.nih.gov/pmc/articles/PMC4533825/

Semantic Segmentation:
https://www.jeremyjordan.me/semantic-segmentation/

How to use sitk filter:
https://www.programmersought.com/article/5062239478/

About the Filter:
https://mevislabdownloads.mevis.de/docs/current/FMEwork/ITK/Documentation/Publish/ModuleReference/itkHausdorffDistanceImageFilter.html

'''
from Data import get_dict_of_paths
import SimpleITK as sitk
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font
import numpy as np
from helpers import get_organ_label, delete_recreate_folder, nifti_image_affine_reader, bb_mm_to_vox, get_bb_coordinates
import nibabel as nib


def calculate_loss_and_accuracy(model, X_test, y_test):
    # Generate generalization metrics
    score = model.evaluate(X_test, y_test, verbose=0)
    print(f'Test loss: {score[0]} / Test accuracy: {score[1]}')


# returns dice coefficent, mean overlap and volume similarity measurements
def calculate_label_overlap_measures(gt_img_path, pred_img_path):
    # load images as int since float is not supported by GetDiceCoefficient()
    gt_img = sitk.ReadImage(gt_img_path, sitk.sitkInt32)
    pred_img = sitk.ReadImage(pred_img_path, sitk.sitkInt32)

    # calculate dice
    dice_filter = sitk.LabelOverlapMeasuresImageFilter()
    dice_filter.Execute(gt_img, pred_img)
    dice = dice_filter.GetDiceCoefficient()
    mean_overlap = dice_filter.GetMeanOverlap()
    volume_similarity = dice_filter.GetVolumeSimilarity()

    print("dice coefficient {}".format(dice))
    print("mean overlap {}".format(mean_overlap))
    print("volume similarity {}".format(volume_similarity))

    return dice, mean_overlap, volume_similarity


# returns hausdorff distance and average hausdorff distance measurements
def calculate_hausdorff_distance(gt_img_path, pred_img_path):
    # load images
    gt_img = sitk.ReadImage(gt_img_path)
    pred_img = sitk.ReadImage(pred_img_path)

    # calculate hausdorff
    hd_filter = sitk.HausdorffDistanceImageFilter()
    hd_filter.Execute(gt_img, pred_img)
    avg_hd = hd_filter.GetAverageHausdorffDistance()
    hd = hd_filter.GetHausdorffDistance()

    print("hausdorff distance {}".format(hd))
    print("average hausdorff distance {}".format(avg_hd))

    return hd, avg_hd


def calculate_mean(results):
    # organize metrics of all patients
    hd = []
    avg_hd = []
    dice = []
    avg_ol = []
    vs = []
    for result in results:
        hd.append(result[1])
        avg_hd.append(result[2])
        dice.append((result[3]))
        avg_ol.append(result[4])
        vs.append(result[5])

    # calculate mean of all metrics
    mean_hd = np.mean(hd)
    mean_avg_hd = np.mean(avg_hd)
    mean_dice = np.mean(dice)
    mean_avg_ol = np.mean(avg_ol)
    mean_vs = np.mean(vs)

    mean = ["mean", mean_hd, mean_avg_hd, mean_dice, mean_avg_ol, mean_vs]
    return mean


def calculate_standard_dv(results):
    # organize metrics of all patients
    hd = []
    avg_hd = []
    dice = []
    avg_ol = []
    vs = []
    for result in results:
        hd.append(result[1])
        avg_hd.append(result[2])
        dice.append((result[3]))
        avg_ol.append(result[4])
        vs.append(result[5])

    # calculate std of all metrics
    std_hd = np.std(hd)
    std_avg_hd = np.std(avg_hd)
    std_dice = np.std(dice)
    std_avg_ol = np.std(avg_ol)
    std_vs = np.std(vs)

    std = ["standard d", std_hd, std_avg_hd, std_dice, std_avg_ol, std_vs]
    return std


def evaluate_predictions(pred_path, gt_path):
    results = []
    std = []

    dict_gt_paths = get_dict_of_paths(gt_path)

    # evaluate every prediction
    for prediction in os.scandir(pred_path):
        # find patient number in file name
        regex = re.compile(r'\d+')
        patient_no = int(regex.search(prediction.name).group(0))
        gt_img_path = dict_gt_paths[patient_no]

        print("Patient Number : {}".format(patient_no))
        hd, avg_hd = calculate_hausdorff_distance(gt_img_path, prediction.path)
        dice, avg_ol, vs = calculate_label_overlap_measures(gt_img_path, prediction.path)
        print("")

        results.append([patient_no, hd, avg_hd, dice, avg_ol, vs])

    mean = calculate_mean(results)
    std = calculate_standard_dv(results)

    return results, mean, std


'''
https://realpython.com/openpyxl-excel-spreadsheets-python/
'''


def create_excel_sheet(organ, save_path):
    # create excel sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = organ

    # create headings and apply style
    headings_style = NamedStyle(
        name="daria",
        font=Font(color='000000', bold=True),
        alignment=Alignment(horizontal='left')
    )
    headings_row = '1'
    headings = ["patient #", "hausdorff dist",
                "Ø hausdorff dist",
                "dice coeff", "Ø overlap",
                "volume similarity"]
    sheet.append(headings)
    for cell in sheet[headings_row]:
        cell.style = headings_style

    # make cells wider
    column_width = 20
    sheet.column_dimensions['A'].width = column_width
    sheet.column_dimensions['B'].width = column_width
    sheet.column_dimensions['C'].width = column_width
    sheet.column_dimensions['D'].width = column_width
    sheet.column_dimensions['E'].width = column_width
    sheet.column_dimensions['F'].width = column_width

    wb.save("{}Evaluation {}.xlsx".format(save_path, organ))


def fill_excel_sheet(save_path_results, save_path_y_test, organ, save_path):
    # open excel sheet
    wb = load_workbook(filename="{}Evaluation {}.xlsx".format(save_path, organ))
    sheet = wb.active

    # get metrics
    results, mean, std = evaluate_predictions(save_path_results, save_path_y_test)

    # write metrics into excel
    number_of_results = len(results)
    number_of_metrics = len(results[0])
    start_row = 2
    start_col = 1
    for i in range(0, number_of_results):
        for j in range(0, number_of_metrics):
            row = start_row + i  # start a new row for every result
            col = start_col + j  # put metrcis into the right column
            # cut all values to 2 decimal places. Except patient number
            metric_value = "{:.2f}".format(results[i][j])
            if j == 0:
                metric_value = "{}".format(results[i][j])
            sheet.cell(column=col, row=row, value=metric_value)

    # write mean and std also into excel
    mean_row = start_row + number_of_results
    std_row = mean_row + 1
    for i in range(0, len(mean)):
        col = start_col + i
        if i == 0:
            mean_value = "{}".format(mean[i])
            std_value = "{}".format(std[i])
        else:
            mean_value = "{:.2f}".format(mean[i])
            std_value = "{:.2f}".format(std[i])
        sheet.cell(column=col, row=mean_row, value=mean_value)
        sheet.cell(column=col, row=std_row, value=std_value)

    wb.save("{}Evaluation {}.xlsx".format(save_path, organ))


def get_dict_of_test_data(input_dict, split):
    result_dict = {}
    # check how many files are in folder and only take the last few ones
    total_files = len(input_dict)
    test = int(total_files * split)
    counter = 0
    for key in sorted(input_dict.keys(), reverse=True):
        if counter == test: break
        result_dict[key] = input_dict[key]
        counter = counter + 1

    return result_dict


def get_segmentation_mask2(img, organ, bb_coords):

    img_arr = img.get_fdata()
    spacing, offset = nifti_image_affine_reader(img)
    # get respective label for the given organ
    organ_label = get_organ_label(organ)

    # create empty (only zeros) segmentation mask with same siza as result_img_arr
    # should be 64,64,64
    result_img_arr = np.zeros((img_arr.shape[0],
                         img_arr.shape[1],
                         img_arr.shape[2]))


    bb_coords_vox = bb_mm_to_vox(bb_coords, spacing, offset)
    print(bb_coords_vox)

    # loop over every voxel in area of interest and create segmentation mask
    for x in range(int(bb_coords_vox[0])-1, int(bb_coords_vox[1])+1):
        for y in range(int(bb_coords_vox[2])-1, int(bb_coords_vox[3])+1):
            for z in range(int(bb_coords_vox[4])-1, int(bb_coords_vox[5])+1):
                # values > thresh will be labeled as segmentation mask
                # result_img_arr should have shape 64,64,64,1
                if img_arr[x][y][z] == organ_label:
                    result_img_arr[x, y, z] = organ_label

    return result_img_arr


def get_segmentation_masks2(dict_files, path_ref_files, save_path, organ, dict_bb_paths):
    delete_recreate_folder(save_path)

    dict_ref_file_paths = get_dict_of_paths(path_ref_files)

    print("get segmentation masks")
    for key in sorted(dict_files.keys()):
        file = dict_files[key]
        bb_path = dict_bb_paths[key]
        bb_coords = get_bb_coordinates(bb_path)

        # get the i-th reference file (patients in ascending order)
        curr_file_path = dict_ref_file_paths[key]

        # check voxel values against treshold and get segmentationmask
        result_img_arr = get_segmentation_mask2(file, organ, bb_coords)

        # save cropped array as nifti file with patient number in name
        ref_file = nib.load(curr_file_path)    # reference file
        result_img = nib.Nifti1Image(result_img_arr, ref_file.affine, ref_file.header)
        nib.save(result_img, '{}seg{}.nii.gz'.format(save_path, key))