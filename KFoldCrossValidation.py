import numpy as np
from sklearn.model_selection import KFold
from SharedMethods import find_patient_no_in_file_name
import os
import tensorflow as tf
from UNet3D.Prepare import prepare
from UNet3D.Train import train
from UNet3D.Apply import apply
from UNet3D.Evaluate import evaluate
from UNet2D.Prepare2D import create_excel_sheet2D
from UNet2D.Train2D import train2D
from UNet2D.Apply2D import apply2D
from UNet2D.Evaluate2D import evaluate2D
import time

from openpyxl import load_workbook
import numpy as np
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl import Workbook
import openpyxl as op

def get_files_in_path(path):
    # count how many files are in SCAN_PATH
    all_patient_numbers = set()
    for file in os.scandir(path):
        patient_no = find_patient_no_in_file_name(file.name)
        all_patient_numbers.add(patient_no)
    amount_patients = len(all_patient_numbers)
    all_patient_numbers_arr = np.zeros(amount_patients)
    i = 0
    for val in all_patient_numbers:
        all_patient_numbers_arr[i] = val
        i = i + 1

    return all_patient_numbers_arr, amount_patients


def run_KfoldCV(SCAN_PATH, GT_BB_PATH, RRF_BB_PATH, GT_SEG_PATH, SAVE_PATH, DIMENSIONS, BATCH, EPOCHS, organs, direction):
    kfold = KFold(5, True, 1)
    data, amount_patients = get_files_in_path(SCAN_PATH)
    parts = kfold.split(data)

    number = 0
    for train_set, test_set in parts:
        print("#{} train set: {} files".format(number, len(train_set)))
        print(train_set)
        print("#{} test set: {} files".format(number, len(test_set)))
        print(test_set)
        print("")

        number = number + 1
        for organ in organs:
            if organ == 'pancreas':
                thresh = 0.3
            else:
                thresh = 0.5

            prepare(SCAN_PATH, GT_BB_PATH, RRF_BB_PATH, GT_SEG_PATH, SAVE_PATH, DIMENSIONS, organ, train_set, test_set)
            create_excel_sheet2D(SAVE_PATH, organ, train_set, test_set)

            # 3D
            start = time.time()
            #train(SAVE_PATH, DIMENSIONS, organ, 0.0, BATCH, EPOCHS)
            end = time.time()
            elapsed_time = end - start

            #apply(SCAN_PATH, RRF_BB_PATH, SAVE_PATH, DIMENSIONS, organ, thresh)
            #evaluate(SAVE_PATH, organ, number, elapsed_time)
            #exit()
            # 2D
            start = time.time()
            train2D(SAVE_PATH, DIMENSIONS, organ, direction, 0.0, BATCH, EPOCHS)
            end = time.time()
            elapsed_time = end - start

            apply2D(SCAN_PATH, RRF_BB_PATH, SAVE_PATH, DIMENSIONS, organ, thresh, direction)
            evaluate2D(SAVE_PATH, organ, number, elapsed_time)



def summarize_eval(SAVE_PATH, ORGAN):
    # create excel sheet
    eval_wb = Workbook()
    eval_sheet = eval_wb.active
    eval_sheet.title = "summarize eval"

    # create headings and apply style
    headings_style = NamedStyle(
        name="daria",
        font=Font(color='000000', bold=True),
        alignment=Alignment(horizontal='left')
    )
    headings_row = '1'
    headings = ["file #", "mean hd",
                "std", "mean avd",
                "std", "mean dice",
                "std", "time"]
    eval_sheet.append(headings)
    for cell in eval_sheet[headings_row]:
        cell.style = headings_style

    # make cells wider
    eval_sheet.column_dimensions['A'].width = 50
    eval_sheet.column_dimensions['B'].width = 20
    eval_sheet.column_dimensions['C'].width = 20
    eval_sheet.column_dimensions['D'].width = 20
    eval_sheet.column_dimensions['E'].width = 20
    eval_sheet.column_dimensions['F'].width = 20
    eval_sheet.column_dimensions['G'].width = 20
    eval_sheet.column_dimensions['H'].width = 20

    mean_dice_row = 20
    mean_standard_d_row = 21
    mean_avd_row = 20
    mean_standard_avd_row = 21
    mean_hd_row = 20
    mean_standard_hd_row = 21
    #mean_asd_row = 20
    #mean_standard_asd_row = 21
    mean_time_row = 22

    relevant_col = 2

    curr_row = 2
    curr_col = 1

    path = SAVE_PATH
    all_dice = []
    all_dice_std = []
    all_hd = []
    all_hd_std = []
    all_avd = []
    all_avd_std = []
    all_time = []
    for file in os.scandir(path):
        found = file.name.find(ORGAN)
        if found is not -1:
            # open
            wb_obj = op.load_workbook(file)
            # get active sheet
            sheet_obj = wb_obj.active

            # read cells
            # HD
            cell_mean_hd = sheet_obj.cell(row=mean_hd_row, column=relevant_col)
            cell_standard_hd = sheet_obj.cell(row=mean_standard_hd_row, column=relevant_col)
            #AVD
            cell_mean_avd = sheet_obj.cell(row=mean_avd_row, column=relevant_col+1)
            cell_standard_avd = sheet_obj.cell(row=mean_standard_avd_row, column=relevant_col+1)
            #DICE
            cell_mean_dice = sheet_obj.cell(row=mean_dice_row, column=relevant_col+2)
            cell_standard_d = sheet_obj.cell(row=mean_standard_d_row, column=relevant_col+2)
            #ASD
            #cell_mean_asd = sheet_obj.cell(row=mean_asd_row, column=relevant_col+3)
            #cell_standard_asd = sheet_obj.cell(row=mean_standard_asd_row, column=relevant_col+3)
            #TIME
            cell_mean_time = sheet_obj.cell(row=mean_time_row, column=relevant_col)


            # get values
            mean_dice = cell_mean_dice.value
            standard_d = cell_standard_d.value
            mean_avd = cell_mean_avd.value
            standard_avd = cell_standard_avd.value
            mean_hd = cell_mean_hd.value
            standard_hd = cell_standard_hd.value
            #mean_asd = cell_mean_asd.value
            #standard_asd = cell_standard_asd.value
            mean_time = cell_mean_time.value

            all_dice.append(mean_dice)
            all_dice_std.append(standard_d)
            all_avd.append(mean_avd)
            all_avd_std.append(standard_avd)
            all_hd.append(mean_hd)
            all_hd_std.append(standard_hd)
            all_time.append(mean_time)

            # write into evaluation summary sheet
            patient_no = find_patient_no_in_file_name(file.name)
            eval_sheet.cell(column=curr_col, row=curr_row, value=patient_no)
            #HD
            eval_sheet.cell(column=curr_col+1, row=curr_row, value=mean_hd)
            eval_sheet.cell(column=curr_col+2, row=curr_row, value=standard_hd)
            #AVD
            eval_sheet.cell(column=curr_col+3, row=curr_row, value=mean_avd)
            eval_sheet.cell(column=curr_col+4, row=curr_row, value=standard_avd)
            #DICE
            eval_sheet.cell(column=curr_col+5, row=curr_row, value=mean_dice)
            eval_sheet.cell(column=curr_col+6, row=curr_row, value=standard_d)
            #ASD
            #eval_sheet.cell(column=curr_col+7, row=curr_row, value=mean_asd)
            #eval_sheet.cell(column=curr_col+8, row=curr_row, value=standard_asd)
            #TIME
            eval_sheet.cell(column=curr_col+7, row=curr_row, value=mean_time)

            curr_row = curr_row+1

    print(all_dice)
    all_dice = list(map(float, all_dice))
    print(all_dice)
    avg_all_dice = np.mean(all_dice)
    print(avg_all_dice)

    all_dice_std = list(map(float, all_dice_std))
    avg_all_dice_std = np.mean(all_dice_std)
    all_avd = list(map(float, all_avd))
    avg_all_avd = np.mean(all_avd)
    all_avd_std = list(map(float, all_avd_std))
    avg_all_avd_std = np.mean(all_avd_std)
    all_hd = list(map(float, all_hd))
    avg_all_hd = np.mean(all_hd)
    all_hd_std = list(map(float, all_hd_std))
    avg_all_hd_std = np.mean(all_hd_std)
    all_time = list(map(float, all_time))
    avg_all_time = np.mean(all_time)

    curr_col = 2
    curr_row = 7

    eval_sheet.cell(column=curr_col, row=curr_row, value=avg_all_hd)
    eval_sheet.cell(column=curr_col+1, row=curr_row, value=avg_all_hd_std)

    eval_sheet.cell(column=curr_col+2, row=curr_row, value=avg_all_avd)
    eval_sheet.cell(column=curr_col+3, row=curr_row, value=avg_all_avd_std)

    eval_sheet.cell(column=curr_col+4, row=curr_row, value=avg_all_dice)
    eval_sheet.cell(column=curr_col+5, row=curr_row, value=avg_all_dice_std)

    eval_sheet.cell(column=curr_col+6, row=curr_row, value=avg_all_time)




    eval_wb.save("{}Evaluation Summary {}.xlsx".format(SAVE_PATH, ORGAN))


def summarize_metrics(SAVE_PATH, metric):
    organs = ['liver', 'left_kidney', 'right_kidney', 'spleen', 'pancreas']

    # create excel sheet
    eval_wb = Workbook()
    eval_sheet = eval_wb.active

    eval_sheet.title = "summarize {}".format(metric)

    # create headings and apply style
    headings_style = NamedStyle(
        name="daria",
        font=Font(color='000000', bold=True),
        alignment=Alignment(horizontal='left')
    )
    headings_row = '1'
    headings = ["file #", "liver 3D",
                "liver 2D", "r. kidney 3D",
                "r. kidney 2D", "l. kidney 3D",
                "l. kidney 2D", "spleen 3D",
                "spleen 2D", "pancreas 3D",
                "pancreas 2D"]
    eval_sheet.append(headings)
    for cell in eval_sheet[headings_row]:
        cell.style = headings_style

    # make cells wider
    #eval_sheet.column_dimensions['A'].width = 50
    eval_sheet.column_dimensions['B'].width = 20
    eval_sheet.column_dimensions['C'].width = 20
    eval_sheet.column_dimensions['D'].width = 20
    eval_sheet.column_dimensions['E'].width = 20
    eval_sheet.column_dimensions['F'].width = 20
    eval_sheet.column_dimensions['G'].width = 20
    eval_sheet.column_dimensions['H'].width = 20
    eval_sheet.column_dimensions['I'].width = 20
    eval_sheet.column_dimensions['J'].width = 20
    eval_sheet.column_dimensions['K'].width = 20

    dice_row = 4
    if metric == "dice":
        dice_col = 4
    elif metric == "hd":
        dice_col = 2
    elif metric == "avd":
        dice_col = 3

    path = SAVE_PATH

    for ORGAN in organs:
        for file in os.scandir(path):
            all_dice = []
            all_patient_no = []
            found = file.name.find(ORGAN)
            if found is not -1:
                # open
                wb_obj = op.load_workbook(file)
                # get active sheet
                sheet_obj = wb_obj.active

                for i in range(0, 16):
                    # read cells

                    cell_dice = sheet_obj.cell(row=dice_row+i, column=dice_col)
                    cell_patient_no = sheet_obj.cell(row=dice_row+i, column=1)

                    # get values
                    dice = cell_dice.value
                    patient_no = cell_patient_no.value
                    all_dice.append(dice)
                    all_patient_no.append(patient_no)
                print(file.name)
                print(all_dice)
                print(all_patient_no)

                # write into evaluation summary sheet
                file_no_1 = file.name.find("1_")
                file_no_2 = file.name.find("2_")
                file_no_3 = file.name.find("3_")
                file_no_4 = file.name.find("4_")
                file_no_5 = file.name.find("5_")

                if file_no_1 is not -1:
                    curr_row = 2
                elif file_no_2 is not -1:
                    curr_row = 18
                elif file_no_3 is not -1:
                    curr_row = 34
                elif file_no_4 is not -1:
                    curr_row = 50
                elif file_no_5 is not -1:
                    curr_row = 66

                switcher = {
                    "liver": 2,
                    "left_kidney": 6,
                    "right_kidney": 4,
                    "spleen": 8,
                    "pancreas": 10
                }
                curr_col = switcher.get(ORGAN)
                found2 = file.name.find("Evaluation2D")
                if found2 is not -1:
                    curr_col = curr_col + 1

                for j in range(0, 16):
                    #DICE
                    eval_sheet.cell(column=curr_col, row=curr_row+j, value=all_dice[j])
                    eval_sheet.cell(column=1, row=curr_row+j, value=all_patient_no[j])


    eval_wb.save("{}{} Summary.xlsx".format(SAVE_PATH, metric))