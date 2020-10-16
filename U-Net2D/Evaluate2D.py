from SharedMethods import get_dict_of_paths, find_patient_no_in_file_name
import SimpleITK as sitk
import os
from openpyxl import load_workbook
import numpy as np
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl import Workbook
import openpyxl as op

def evaluate(SAVE_PATH, ORGAN, ROUND):

    # open excel sheet
    wb = load_workbook(filename="{}2DEvaluation {}.xlsx".format(SAVE_PATH, ORGAN))
    sheet = wb.active

    # get metrics
    path_results_orig = "{}results/orig/".format(SAVE_PATH)
    path_y_test_orig = "{}ytest/orig/".format(SAVE_PATH)
    results, mean, std = evaluate_predictions(path_results_orig, path_y_test_orig)

    # write metrics into excel
    number_of_results = len(results)
    number_of_metrics = len(results[0])
    start_row = 4
    start_col = 1
    for i in range(0, number_of_results):
        for j in range(0, number_of_metrics):
            row = start_row + i  # start a new row for every result
            col = start_col + j  # put metrcis into the right column
            # cut all values to 2 decimal places. Except patient number
            metric_value = "{:.2f}".format(results[i][j])
            if j == 0:
                metric_value = "{}".format(results[i][j])
            sheet.cell(column=col, row=row, value=metric_value)

    # write mean and std also into excel
    mean_row = start_row + number_of_results
    std_row = mean_row + 1
    for i in range(0, len(mean)):
        col = start_col + i
        if i == 0:
            mean_value = "{}".format(mean[i])
            std_value = "{}".format(std[i])
        else:
            mean_value = "{:.2f}".format(mean[i])
            std_value = "{:.2f}".format(std[i])
        sheet.cell(column=col, row=mean_row, value=mean_value)
        sheet.cell(column=col, row=std_row, value=std_value)

    wb.save("{}eval/{}_2DEvaluation {}.xlsx".format(SAVE_PATH, ROUND, ORGAN))
